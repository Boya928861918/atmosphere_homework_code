import numpy as np
import netCDF4 as nc
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeat
import openpyxl
from cartopy.mpl.gridliner import LONGITUDE_FORMATTER, LATITUDE_FORMATTER
from numpy.core.arrayprint import format_float_positional

wbu = openpyxl.load_workbook('u.xlsx')
wbv = openpyxl.load_workbook('v.xlsx')
wbr = openpyxl.load_workbook('R.xlsx')
wsu = wbu['u']
wsv = wbv['v']
wsr = wbr['Sheet1']
col = 17
row = 17
u, v = [], []
for i in range(col):
    u0 = []
    v0 = []
    for j in range(row):
        u0.append(wsu.cell(row = i+2, column = j+2).value)
        v0.append(wsv.cell(row = i+2, column = j+2).value)
    u.append(u0)
    v.append(v0)
r = []
for i in range(17):
    r0 = []
    for j in range(17):
        r0.append(wsr.cell(row = i+1, column = j+1).value)
    r.append(r0)
x, y = np.arange(-2,2.25,0.25), np.arange(-2,2.25,0.25)

# 创建底图
fig, ax = plt.subplots(figsize=(12, 12), dpi=100)
# ax.quiver(x, y, u, v, scale = 150)
print(u)
print('------------------------------------------------')
print(v)
ax.quiver(x, y, u, v, scale = 150)

b = plt.contour(x,y,r, levels = 20, colors='k')
plt.clabel(b, inline=True, fontsize=10)
plt.savefig('uv.png')